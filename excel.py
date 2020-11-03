import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelPainter:

    def __init__(self, filename, startingRow, startingColumn):
        # Open an xlsx for reading
        self.wb = load_workbook(filename=filename)
        # Get the current Active Sheet
        self.ws = self.wb.get_active_sheet()
        # get the row index for the xlsx
        self.redFill = PatternFill(start_color='FFFF0000',
                                   end_color='FFFF0000',
                                   fill_type='solid')
        self.fileName = filename
        self.startingRow = startingRow
        self.startingColumn = startingColumn
        self.startingTimeRow = 17

    def saveValue(self, rowNumber, columnNumber, value):
        self.ws.cell(row=(rowNumber), column=(columnNumber)).value = value

    def saveFile(self):
        self.wb.save(self.fileName)

    def paintCell(self, rowNumber, columnNumber):
        self.ws.cell(row=(rowNumber + self.startingRow), column=(columnNumber + self.startingColumn)).fill = self.redFill

    def paintCells(self, startingColumn, endingColumn, row):
        for column in range(startingColumn, endingColumn):
            self.paintCell(row,column)